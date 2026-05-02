import io
from collections import defaultdict
from datetime import date, datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

DATE_FMT = "DD-MM-YYYY"

ALLOWED_FILTER_NAMES = {
    "Тип оборудования",
    "Категория",
    "Модель",
    "Серийный номер",
    "Инвентарный номер",
    "Сетевое имя",
    "Ответственное лицо",
    "Статус",
    "Адрес",
    "Помещение",
    "Подразделение",
    "Дата принятия от",
    "Дата принятия до",
}

def _as_date(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return None


def _dt(v):
    """Для сортировок: date/datetime -> datetime; None -> datetime.min"""
    if v is None:
        return datetime.min
    if isinstance(v, datetime):
        return v
    if isinstance(v, date):
        return datetime(v.year, v.month, v.day)
    return datetime.min


def build_grouped_report(equipment_items) -> bytes:
    """
    Группировка
    equipment_items: список Equipment с полями:
      - e.model / e.name
      - e.inventory_number
      - e.accepted_date / e.accepted_at
      - e.type.type_name / e.type.name
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Группировка"

    header_font = Font(bold=True)
    group_font = Font(bold=True)
    base_font = Font(bold=False)

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    header_fill = PatternFill("solid", fgColor="D9E1F2")
    group_fill = PatternFill("solid", fgColor="F2F2F2")

    widths = {"A": 40, "B": 22, "C": 18, "D": 25, "E": 16}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    headers = [
        "Наименование / модель",
        "Инвентарный номер",
        "Дата принятия к учету",
        "Тип оборудования",
        "Итого за период",
    ]
    ws.append(headers)

    for col in range(1, 6):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.alignment = center
        cell.fill = header_fill
        cell.border = border
    ws.row_dimensions[1].height = 22

    rows = []
    for e in equipment_items:
        t = getattr(e, "type", None)
        type_name = getattr(t, "type_name", "") or getattr(t, "name", "") or ""
        model = getattr(e, "model", "") or getattr(e, "name", "") or ""
        inv = getattr(e, "inventory_number", "") or ""
        accepted_raw = getattr(e, "accepted_date", None) or getattr(e, "accepted_at", None)
        accepted = _as_date(accepted_raw)

        rows.append((type_name, model, inv, accepted))

    rows.sort(key=lambda r: (r[0], _dt(r[3]), r[2]))

    grouped = defaultdict(list)
    for type_name, model, inv, accepted in rows:
        grouped[type_name].append((model, inv, accepted, type_name))

    current_row = 2
    for type_name, items in grouped.items():
        ws.append([type_name, "", "", "", len(items)])
        for col in range(1, 6):
            cell = ws.cell(row=current_row, column=col)
            cell.font = group_font
            cell.alignment = left if col == 1 else center
            cell.fill = group_fill
            cell.border = border
        ws.row_dimensions[current_row].height = 18
        current_row += 1

        items.sort(key=lambda x: (_dt(x[2]), x[1]))

        for model, inv, accepted, tname in items:
            ws.append([model, inv, accepted, tname, ""])

            # формат даты в колонке C
            date_cell = ws.cell(row=current_row, column=3)
            date_cell.number_format = DATE_FMT

            for col in range(1, 6):
                cell = ws.cell(row=current_row, column=col)
                cell.font = base_font
                cell.alignment = left if col in (1, 4) else center
                cell.border = border

            ws.row_dimensions[current_row].height = 18
            current_row += 1

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:E{max(1, ws.max_row)}"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_complex_report_all_categories(
    blocks: list[tuple[str, list[tuple[str, int, int]]]],
) -> bytes:
    """
    blocks: [(category_name, [(type_name, total_all, total_period), ...]),...]
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Сложносоставной"

    header_font = Font(bold=True)
    bold_font = Font(bold=True)

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    header_fill = PatternFill("solid", fgColor="D9E1F2")
    total_fill = PatternFill("solid", fgColor="F2F2F2")

    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 30

    ws.append(["Наименование", "Всего", "Поступившие за указанный период"])
    for col in range(1, 4):
        c = ws.cell(1, col)
        c.font = header_font
        c.alignment = center
        c.fill = header_fill
        c.border = border
    ws.row_dimensions[1].height = 22

    r = 2
    for category_name, rows in blocks:
        if not rows:
            continue

        total_all = sum(x[1] for x in rows)
        total_period = sum(x[2] for x in rows)

        ws.append([f"{category_name}, всего", total_all, total_period])
        for col in range(1, 4):
            c = ws.cell(r, col)
            c.font = bold_font
            c.alignment = left if col == 1 else center
            c.fill = total_fill
            c.border = border
        r += 1

        for type_name, all_cnt, period_cnt in rows:
            ws.append([f"из них {type_name.lower()}", all_cnt, period_cnt])
            for col in range(1, 4):
                c = ws.cell(r, col)
                c.alignment = left if col == 1 else center
                c.border = border
            r += 1

        ws.append(["", "", ""])
        r += 1

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

def build_simple_table_report(equipment_items, filters: list[dict] | None = None) -> bytes:
    """
    Report type 3:
    Вверху (опционально) таблица "Примененные фильтры"
    Потом таблица оборудования как на скрине.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Таблица"

    header_font = Font(bold=True)
    title_font = Font(bold=True, size=14)

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    header_fill = PatternFill("solid", fgColor="D9E1F2")
    section_fill = PatternFill("solid", fgColor="F2F2F2")

    # Excel date format (как на скрине: 02.02.2023)
    DATE_FMT_LOCAL = "DD.MM.YYYY"

    def _latest_status(e):
        statuses = list(getattr(e, "statuses", None) or [])
        if not statuses:
            return None
        return sorted(statuses, key=lambda s: s.status_change_date, reverse=True)[0]

    # --- 1) подготовка фильтров (только whitelist) ---
    clean_filters: list[tuple[str, str]] = []
    for f in (filters or []):
        name = (f.get("name") or "").strip()
        value = (f.get("value") or "").strip()
        if name in ALLOWED_FILTER_NAMES:
            clean_filters.append((name, value))

    row = 1

    # --- 2) таблица фильтров (если есть) ---
    if clean_filters:
        ws.cell(row=row, column=1, value="Примененные фильтры").font = title_font
        row += 1

        ws.append(["Фильтр", "Значение"])
        for col in range(1, 3):
            c = ws.cell(row=row, column=col)
            c.font = header_font
            c.alignment = center
            c.fill = header_fill
            c.border = border
        ws.row_dimensions[row].height = 20
        row += 1

        for name, value in clean_filters:
            ws.append([name, value])
            for col in range(1, 3):
                c = ws.cell(row=row, column=col)
                c.alignment = left
                c.border = border
            ws.row_dimensions[row].height = 18
            row += 1

        row += 1  # пустая строка

    # --- 3) заголовок секции "Оборудование" ---
    ws.cell(row=row, column=1, value="Оборудование").font = title_font
    row += 1

    # --- 4) основная таблица ---
    headers = [
        "Тип оборудования",
        "Модель оборудования",
        "Серийный номер",
        "Инвентарный номер",
        "Сетевое имя",
        "ФИО ответственного",
        "Подразделение",
        "Статус оборудования",
        "Адрес корпуса",
        "Помещение",
        "Дата принятия к учету",
    ]

    ws.append(headers)
    header_row = row
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=header_row, column=col)
        c.font = header_font
        c.alignment = center
        c.fill = header_fill
        c.border = border
    ws.row_dimensions[header_row].height = 22
    row += 1

    # ширины (примерно как на скрине)
    widths = {
        "A": 22, "B": 20, "C": 18, "D": 18, "E": 16,
        "F": 26, "G": 18, "H": 18, "I": 30, "J": 18, "K": 18,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    for e in equipment_items:
        t = getattr(e, "type", None)
        type_name = getattr(t, "type_name", "") or getattr(t, "name", "") or ""

        latest = _latest_status(e)

        # responsible + office
        fio = None
        office = None
        status_name = None
        building_addr = None
        room_display = None

        if latest is not None:
            # status
            st = getattr(latest, "status_type", None)
            status_name = getattr(st, "status_type_name", None) if st else None

            # responsible
            ru = getattr(latest, "responsible_user", None)
            if ru is not None:
                fio = f"{ru.last_name} {ru.first_name} {ru.paternity}".strip()
                off = getattr(ru, "office", None)
                office = getattr(off, "office_name", None) if off else None

            # room + building + room_type => "112 (Аудитория)"
            room = getattr(latest, "room", None)
            if room is not None:
                rname = getattr(room, "name", None)
                rtype = getattr(getattr(room, "room_type", None), "room_type", None)
                if rname and rtype:
                    room_display = f"{rname} ({rtype})"
                else:
                    room_display = rname

                b = getattr(room, "building", None)
                building_addr = getattr(b, "building_address", None) if b else None

        accepted = _as_date(getattr(e, "accepted_date", None))

        ws.append(
            [
                type_name,
                getattr(e, "model", None),
                getattr(e, "serial_number", None),
                getattr(e, "inventory_number", None),
                getattr(e, "network_name", None),
                fio,
                office,
                status_name,
                building_addr,
                room_display,
                accepted,
            ]
        )

        data_row = row
        # дата в колонке K
        date_cell = ws.cell(row=data_row, column=11)
        date_cell.number_format = DATE_FMT_LOCAL

        for col in range(1, len(headers) + 1):
            c = ws.cell(row=data_row, column=col)
            c.border = border
            c.alignment = left if col in (1, 2, 6, 7, 9, 10) else center

        ws.row_dimensions[data_row].height = 18
        row += 1

    # Freeze/Filter от заголовка таблицы оборудования
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1).coordinate
    ws.auto_filter.ref = f"A{header_row}:K{max(header_row, ws.max_row)}"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()